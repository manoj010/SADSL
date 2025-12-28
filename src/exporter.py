from openpyxl import Workbook

def export_to_excel(
    suspicious_logs,
    top_ips,
    targeted_urls,
    output_file="report.xlsx"
):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Suspicious Events"

    ws1.append([
        "Line No", "Timestamp", "IP",
        "Rule", "Severity", "Action", "Status"
    ])

    for e in suspicious_logs:
        ws1.append([
            e["line_no"],
            e["timestamp"],
            e["ip"],
            e["rule"],
            e["severity"],
            e["action"],
            e["status"]
        ])

    ws2 = wb.create_sheet(title="Top Suspicious IPs")
    ws2.append(["IP Address", "Event Count"])

    for ip, count in top_ips:
        ws2.append([ip, count])

    ws3 = wb.create_sheet(title="Targeted URLs")
    ws3.append(["URL", "Count"])

    for url, count in targeted_urls:
        ws3.append([url, count])

    wb.save(output_file)